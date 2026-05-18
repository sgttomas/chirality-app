#!/usr/bin/env python3
"""
render_table_xlsx.py
Deterministically render one PDF2MD `table_data` block to a single .xlsx workbook.

Usage:
    python3 render_table_xlsx.py --table-data-json path/to/td.json \
        --output-xlsx out/tables/MWK_1956_p0003_tbl01_creep-strength.xlsx \
        --caption "Creep strength ratios" --doc-stem MWK_1956 --page 3 --ordinal 1

Inputs:
    --table-data-json  Path to a JSON file containing the `table_data` block
                       (the value of an asset's "table_data" field, NOT the
                       full page asset JSON). May also be `-` to read stdin.
    --output-xlsx      Output .xlsx path. Parent must exist.
    --caption          Table caption (rendered in the provenance notes block).
    --doc-stem         Document stem (rendered in the provenance notes block).
    --page             1-indexed page number.
    --ordinal          1-indexed table ordinal within the page.
    --slug             Optional slug; used to name the worksheet. Defaults to
                       "Table" (sheet names are length-capped to 31 chars).

Determinism contract:
    - `wb.properties.created` and `modified` are pinned to 1970-01-01T00:00:00Z.
    - `wb.properties.creator` / `lastModifiedBy` are pinned to empty strings.
    - No reliance on system time, hostname, or random IDs.
    - Column widths are derived from max-content-length of the rendered cells
      (deterministic; no autosize-after-save behavior).
    - Re-running with byte-identical input produces a byte-identical output.

Schema:
    The block conforms to `pdf2md-table/v1` as documented in
    skills/pdf2md-page-assets/SKILL.md. Structural validation runs here as a
    last-line check; materialize_page_assets.py also pre-validates.
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.cell.cell import Cell
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - environment guard
    raise SystemExit("ERROR: openpyxl is required for table XLSX rendering") from exc


SCHEMA_VERSION = "pdf2md-table/v1"
VALID_TYPES = frozenset({"text", "number", "fraction", "missing", "formula", "boolean"})
EPOCH = datetime(1970, 1, 1, tzinfo=timezone.utc)
HEADER_FILL = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
HEADER_FONT = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Render a pdf2md table_data block to deterministic XLSX.")
    parser.add_argument("--table-data-json", required=True, help="Path to JSON, or '-' for stdin.")
    parser.add_argument("--output-xlsx", required=True)
    parser.add_argument("--caption", default="")
    parser.add_argument("--doc-stem", required=True)
    parser.add_argument("--page", required=True, type=int)
    parser.add_argument("--ordinal", required=True, type=int)
    parser.add_argument("--slug", default="Table")
    return parser.parse_args()


def load_table_data(arg: str) -> dict[str, Any]:
    if arg == "-":
        return json.loads(sys.stdin.read())
    path = Path(arg)
    if not path.is_file():
        raise SystemExit(f"ERROR: table-data-json not found: {arg}")
    return json.loads(path.read_text(encoding="utf-8"))


def validate_table_data(td: dict[str, Any]) -> None:
    """Structural validation matching SKILL.md's pdf2md-table/v1 schema.

    Raises SystemExit(2) with a contract-pointing error on any defect.
    """
    if not isinstance(td, dict):
        raise SystemExit("ERROR: table_data must be a JSON object (see SKILL.md pdf2md-table/v1)")

    sv = td.get("schema_version")
    if sv != SCHEMA_VERSION:
        raise SystemExit(f"ERROR: table_data.schema_version must be '{SCHEMA_VERSION}', got {sv!r}")

    rows = td.get("rows")
    if not isinstance(rows, list) or not rows:
        raise SystemExit("ERROR: table_data.rows must be a non-empty array")

    header_rows = td.get("header_rows")
    if not isinstance(header_rows, int) or header_rows < 0 or header_rows > len(rows):
        raise SystemExit(f"ERROR: table_data.header_rows must be int in [0,{len(rows)}], got {header_rows!r}")

    section_dividers = td.get("section_dividers", [])
    if not isinstance(section_dividers, list):
        raise SystemExit("ERROR: table_data.section_dividers must be an array")
    for idx in section_dividers:
        if not isinstance(idx, int) or idx < 0 or idx >= len(rows):
            raise SystemExit(f"ERROR: section_dividers index out of range: {idx!r}")

    footnotes = td.get("footnotes", [])
    if not isinstance(footnotes, list):
        raise SystemExit("ERROR: table_data.footnotes must be an array")
    footnote_markers: set[str] = set()
    for fn in footnotes:
        if not isinstance(fn, dict) or "marker" not in fn or "text" not in fn:
            raise SystemExit("ERROR: each footnote must be an object with 'marker' and 'text'")
        if not isinstance(fn["marker"], str) or not isinstance(fn["text"], str):
            raise SystemExit("ERROR: footnote 'marker' and 'text' must be strings")
        footnote_markers.add(fn["marker"])

    continuation = td.get("continuation_of")
    if continuation is not None:
        if not isinstance(continuation, dict):
            raise SystemExit("ERROR: continuation_of must be null or an object")
        for key in ("doc_stem", "page", "tbl_ordinal"):
            if key not in continuation:
                raise SystemExit(f"ERROR: continuation_of missing key '{key}'")
        if not isinstance(continuation["doc_stem"], str):
            raise SystemExit("ERROR: continuation_of.doc_stem must be a string")
        if not isinstance(continuation["page"], int) or not isinstance(continuation["tbl_ordinal"], int):
            raise SystemExit("ERROR: continuation_of.page and tbl_ordinal must be integers")

    for r_idx, row in enumerate(rows):
        if not isinstance(row, dict) or not isinstance(row.get("cells"), list):
            raise SystemExit(f"ERROR: rows[{r_idx}] must be an object with a 'cells' array")
        for c_idx, cell in enumerate(row["cells"]):
            if not isinstance(cell, dict):
                raise SystemExit(f"ERROR: rows[{r_idx}].cells[{c_idx}] must be an object")
            if "value" not in cell:
                raise SystemExit(f"ERROR: rows[{r_idx}].cells[{c_idx}] missing 'value'")
            for span_key in ("row_span", "col_span"):
                if span_key in cell:
                    span = cell[span_key]
                    if not isinstance(span, int) or span < 1:
                        raise SystemExit(f"ERROR: rows[{r_idx}].cells[{c_idx}].{span_key} must be int >= 1")
            if "is_header" in cell and not isinstance(cell["is_header"], bool):
                raise SystemExit(f"ERROR: rows[{r_idx}].cells[{c_idx}].is_header must be boolean")
            if "type" in cell and cell["type"] not in VALID_TYPES:
                raise SystemExit(
                    f"ERROR: rows[{r_idx}].cells[{c_idx}].type must be one of {sorted(VALID_TYPES)}, got {cell['type']!r}"
                )
            markers = cell.get("footnote_markers", [])
            if not isinstance(markers, list):
                raise SystemExit(f"ERROR: rows[{r_idx}].cells[{c_idx}].footnote_markers must be an array")
            for m in markers:
                if m not in footnote_markers:
                    raise SystemExit(
                        f"ERROR: rows[{r_idx}].cells[{c_idx}].footnote_markers contains undeclared marker {m!r}"
                    )

    # row_span may not cross a section_dividers boundary.
    divider_set = set(section_dividers)
    for r_idx, row in enumerate(rows):
        for c_idx, cell in enumerate(row["cells"]):
            span = cell.get("row_span", 1)
            if span > 1:
                for k in range(1, span):
                    if (r_idx + k) in divider_set:
                        raise SystemExit(
                            f"ERROR: rows[{r_idx}].cells[{c_idx}].row_span={span} crosses section_divider at row {r_idx + k}"
                        )


def materialize_grid(rows: list[dict[str, Any]]) -> list[list[dict[str, Any] | None]]:
    """Expand the row-major cell list (which omits span-covered slots) into a
    fully-populated 2D grid. Span-covered slots are filled with None.

    Returns a list-of-lists indexed [row][col]. The grid's width is the
    maximum column extent computed across the row sequence.
    """
    grid: list[list[dict[str, Any] | None]] = []
    occupied: dict[tuple[int, int], bool] = {}

    for r_idx, row in enumerate(rows):
        while len(grid) <= r_idx:
            grid.append([])
        c_out = 0
        for cell in row["cells"]:
            while occupied.get((r_idx, c_out)):
                c_out += 1
                while len(grid[r_idx]) <= c_out:
                    grid[r_idx].append(None)
            row_span = cell.get("row_span", 1)
            col_span = cell.get("col_span", 1)
            while len(grid[r_idx]) <= c_out + col_span - 1:
                grid[r_idx].append(None)
            grid[r_idx][c_out] = cell
            for dr in range(row_span):
                for dc in range(col_span):
                    if dr == 0 and dc == 0:
                        continue
                    rr = r_idx + dr
                    cc = c_out + dc
                    while len(grid) <= rr:
                        grid.append([])
                    while len(grid[rr]) <= cc:
                        grid[rr].append(None)
                    occupied[(rr, cc)] = True
            c_out += col_span

    max_cols = max((len(r) for r in grid), default=0)
    for row in grid:
        while len(row) < max_cols:
            row.append(None)
    return grid


def _set_typed_value(cell: Cell, table_cell: dict[str, Any]) -> None:
    t = table_cell.get("type", "text")
    value = table_cell.get("value")
    if t == "missing":
        cell.value = None
    elif t == "boolean":
        cell.value = bool(value) if value is not None else None
    elif t == "number":
        cell.value = value
    elif t in ("text", "fraction", "formula"):
        cell.value = "" if value is None else str(value)
    else:  # pragma: no cover - validate_table_data guards
        cell.value = value


def _attach_footnote_comment(cell: Cell, markers: list[str], footnote_map: dict[str, str]) -> None:
    if not markers:
        return
    lines = [f"{m}: {footnote_map[m]}" for m in markers if m in footnote_map]
    if not lines:
        return
    cell.comment = Comment("\n".join(lines), "pdf2md")


def _column_widths(grid: list[list[dict[str, Any] | None]]) -> list[int]:
    if not grid:
        return []
    n_cols = len(grid[0])
    widths = [8] * n_cols
    for row in grid:
        for c_idx, td_cell in enumerate(row):
            if td_cell is None:
                continue
            v = td_cell.get("value")
            raw = td_cell.get("raw")
            text = raw if isinstance(raw, str) and raw else ("" if v is None else str(v))
            unit = td_cell.get("unit", "")
            if unit:
                text = f"{text} {unit}"
            # +2 padding; cap at 60 to avoid runaway widths
            widths[c_idx] = max(widths[c_idx], min(len(text) + 2, 60))
    return widths


def _sheet_name(slug: str) -> str:
    safe = (slug or "Table").strip() or "Table"
    # Excel sheet names: max 31 chars, none of: : \ / ? * [ ]
    bad = set(':\\/?*[]')
    cleaned = "".join("_" if ch in bad else ch for ch in safe)
    return cleaned[:31] or "Table"


def render(
    table_data: dict[str, Any],
    output_path: Path,
    *,
    caption: str,
    doc_stem: str,
    page: int,
    ordinal: int,
    slug: str,
) -> None:
    validate_table_data(table_data)

    wb = Workbook()
    # Determinism: pin every timestamp + identity field openpyxl emits.
    wb.properties.created = EPOCH
    wb.properties.modified = EPOCH
    wb.properties.creator = ""
    wb.properties.lastModifiedBy = ""
    wb.properties.title = caption or ""
    wb.properties.description = f"{doc_stem} p{page:04d} tbl{ordinal:02d}"

    ws = wb.active
    ws.title = _sheet_name(slug)

    rows = table_data["rows"]
    header_rows = table_data["header_rows"]
    footnotes = table_data.get("footnotes", [])
    footnote_map = {fn["marker"]: fn["text"] for fn in footnotes}
    grid = materialize_grid(rows)

    # Write cells (skip None-slots covered by span).
    for r_idx, row in enumerate(grid):
        excel_row = r_idx + 1
        for c_idx, td_cell in enumerate(row):
            if td_cell is None:
                continue
            excel_col = c_idx + 1
            cell = ws.cell(row=excel_row, column=excel_col)
            _set_typed_value(cell, td_cell)
            is_header = td_cell.get("is_header") or (r_idx < header_rows)
            if is_header:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            cell.alignment = WRAP
            _attach_footnote_comment(cell, td_cell.get("footnote_markers", []), footnote_map)

    # Apply merges.
    for r_idx, row in enumerate(rows):
        c_out = 0
        # Re-derive the actual layout positions per materialize_grid's logic.
        occupied: dict[tuple[int, int], bool] = {}
        # Replay the occupancy map up to this row index.
        for rr in range(r_idx):
            cc = 0
            for cell in rows[rr]["cells"]:
                while occupied.get((rr, cc)):
                    cc += 1
                row_span = cell.get("row_span", 1)
                col_span = cell.get("col_span", 1)
                for dr in range(row_span):
                    for dc in range(col_span):
                        if dr == 0 and dc == 0:
                            continue
                        occupied[(rr + dr, cc + dc)] = True
                cc += col_span
        for cell in row["cells"]:
            while occupied.get((r_idx, c_out)):
                c_out += 1
            row_span = cell.get("row_span", 1)
            col_span = cell.get("col_span", 1)
            if row_span > 1 or col_span > 1:
                ws.merge_cells(
                    start_row=r_idx + 1,
                    start_column=c_out + 1,
                    end_row=r_idx + row_span,
                    end_column=c_out + col_span,
                )
            c_out += col_span

    # Column widths.
    widths = _column_widths(grid)
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # Notes block beneath the data: caption / footnotes / provenance.
    next_row = len(grid) + 2
    notes_col_letter = get_column_letter(1)
    notes: list[str] = []
    if caption:
        notes.append(f"Caption: {caption}")
    for fn in footnotes:
        notes.append(f"Footnote {fn['marker']}: {fn['text']}")
    cont = table_data.get("continuation_of")
    if cont:
        notes.append(
            f"Continuation of {cont['doc_stem']} page {cont['page']} table {cont['tbl_ordinal']}"
        )
    notes.append(f"Extracted from {doc_stem} page {page}, table {ordinal}.")
    for offset, line in enumerate(notes):
        ws[f"{notes_col_letter}{next_row + offset}"] = line

    wb.save(str(output_path))


def main() -> int:
    args = parse_args()
    output_path = Path(args.output_xlsx)
    if not output_path.parent.is_dir():
        print(f"ERROR: output parent directory does not exist: {output_path.parent}", file=sys.stderr)
        return 2

    table_data = load_table_data(args.table_data_json)
    render(
        table_data,
        output_path,
        caption=args.caption,
        doc_stem=args.doc_stem,
        page=args.page,
        ordinal=args.ordinal,
        slug=args.slug,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
