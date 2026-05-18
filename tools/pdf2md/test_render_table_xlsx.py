#!/usr/bin/env python3
"""
Unit tests for render_table_xlsx.py.

Covers:
  - structural validation (every defect class the renderer's
    `validate_table_data` rejects)
  - grid materialization with row/col spans
  - cell typing (text / number / fraction / missing / formula / boolean)
  - header styling (declared `is_header` and the implicit `header_rows` band)
  - footnote-comment attachment
  - byte-for-byte determinism (same input → same output bytes)
  - openpyxl round-trip readback (values + merges + comments survive)

Run:
    python3 -m pytest tools/pdf2md/test_render_table_xlsx.py -v
"""

from __future__ import annotations

import hashlib
import importlib.util
import json
import shutil
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


_HERE = Path(__file__).resolve().parent
_SPEC = importlib.util.spec_from_file_location("render_table_xlsx", _HERE / "render_table_xlsx.py")
_MOD = importlib.util.module_from_spec(_SPEC)
sys.modules["render_table_xlsx"] = _MOD
_SPEC.loader.exec_module(_MOD)


def _td(**overrides) -> dict:
    """Minimal valid table_data block — single header row + one data row."""
    base = {
        "schema_version": "pdf2md-table/v1",
        "header_rows": 1,
        "section_dividers": [],
        "continuation_of": None,
        "footnotes": [],
        "rows": [
            {"cells": [{"value": "A", "is_header": True}, {"value": "B", "is_header": True}]},
            {"cells": [{"value": "x"}, {"value": 1, "type": "number"}]},
        ],
    }
    base.update(overrides)
    return base


def _render(td: dict, out: Path, **kw) -> None:
    defaults = dict(caption="Demo", doc_stem="DOC", page=1, ordinal=1, slug="t")
    defaults.update(kw)
    _MOD.render(td, out, **defaults)


class Validation(unittest.TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="render_xlsx_"))

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def _expect_fail(self, td: dict, needle: str):
        with self.assertRaises(SystemExit) as ctx:
            _MOD.validate_table_data(td)
        self.assertIn(needle, str(ctx.exception))

    def test_happy_path_valid(self):
        _MOD.validate_table_data(_td())

    def test_not_a_dict_rejected(self):
        self._expect_fail([], "must be a JSON object")

    def test_wrong_schema_version_rejected(self):
        self._expect_fail(_td(schema_version="pdf2md-table/v0"), "schema_version")

    def test_rows_empty_rejected(self):
        self._expect_fail(_td(rows=[]), "non-empty array")

    def test_header_rows_out_of_range(self):
        self._expect_fail(_td(header_rows=5), "header_rows")

    def test_header_rows_negative(self):
        self._expect_fail(_td(header_rows=-1), "header_rows")

    def test_section_dividers_out_of_range(self):
        self._expect_fail(_td(section_dividers=[7]), "section_dividers")

    def test_unknown_cell_type_rejected(self):
        bad = _td()
        bad["rows"][1]["cells"][0]["type"] = "currency"
        self._expect_fail(bad, ".type must be one of")

    def test_row_span_zero_rejected(self):
        bad = _td()
        bad["rows"][0]["cells"][0]["row_span"] = 0
        self._expect_fail(bad, "row_span")

    def test_footnote_marker_not_declared_rejected(self):
        bad = _td()
        bad["rows"][1]["cells"][0]["footnote_markers"] = ["*"]
        self._expect_fail(bad, "undeclared marker")

    def test_row_span_crossing_section_divider_rejected(self):
        td = {
            "schema_version": "pdf2md-table/v1",
            "header_rows": 0,
            "section_dividers": [1],
            "continuation_of": None,
            "footnotes": [],
            "rows": [
                {"cells": [{"value": "a", "row_span": 2}]},
                {"cells": []},
            ],
        }
        self._expect_fail(td, "crosses section_divider")

    def test_continuation_of_missing_key_rejected(self):
        bad = _td(continuation_of={"doc_stem": "X", "page": 1})  # missing tbl_ordinal
        self._expect_fail(bad, "continuation_of")

    def test_footnote_text_must_be_string(self):
        bad = _td(footnotes=[{"marker": "*", "text": 123}])
        self._expect_fail(bad, "must be strings")

    def test_cell_missing_value_rejected(self):
        bad = _td()
        bad["rows"][1]["cells"][0] = {"type": "text"}  # no 'value'
        self._expect_fail(bad, "missing 'value'")


class GridMaterialization(unittest.TestCase):
    def test_no_spans_flat_grid(self):
        rows = [{"cells": [{"value": "a"}, {"value": "b"}]}]
        grid = _MOD.materialize_grid(rows)
        self.assertEqual(len(grid), 1)
        self.assertEqual([c["value"] for c in grid[0]], ["a", "b"])

    def test_col_span_expands_width(self):
        rows = [
            {"cells": [{"value": "header", "col_span": 2}]},
            {"cells": [{"value": "a"}, {"value": "b"}]},
        ]
        grid = _MOD.materialize_grid(rows)
        self.assertEqual(len(grid[0]), 2)
        self.assertEqual(grid[0][0]["value"], "header")
        self.assertIsNone(grid[0][1])  # span-covered

    def test_row_span_blocks_subsequent_row(self):
        rows = [
            {"cells": [{"value": "tall", "row_span": 2}, {"value": "x"}]},
            {"cells": [{"value": "y"}]},
        ]
        grid = _MOD.materialize_grid(rows)
        self.assertEqual(grid[0][0]["value"], "tall")
        self.assertIsNone(grid[1][0])  # covered by row_span
        self.assertEqual(grid[1][1]["value"], "y")


class RenderingRoundTrip(unittest.TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="render_xlsx_"))

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_simple_table_values_survive(self):
        out = self.tmp / "t.xlsx"
        _render(_td(), out)
        wb = load_workbook(str(out))
        ws = wb.active
        self.assertEqual(ws.cell(1, 1).value, "A")
        self.assertEqual(ws.cell(2, 1).value, "x")
        self.assertEqual(ws.cell(2, 2).value, 1)

    def test_header_cells_bold(self):
        out = self.tmp / "t.xlsx"
        _render(_td(), out)
        wb = load_workbook(str(out))
        ws = wb.active
        self.assertTrue(ws.cell(1, 1).font.bold)
        self.assertFalse(bool(ws.cell(2, 1).font.bold))

    def test_merge_recorded(self):
        td = {
            "schema_version": "pdf2md-table/v1",
            "header_rows": 2,
            "section_dividers": [],
            "continuation_of": None,
            "footnotes": [],
            "rows": [
                {"cells": [
                    {"value": "Size", "row_span": 2, "is_header": True},
                    {"value": "Rating", "col_span": 2, "is_header": True},
                ]},
                {"cells": [
                    {"value": "psig", "is_header": True},
                    {"value": "°F", "is_header": True},
                ]},
                {"cells": [{"value": "1-1/2"}, {"value": 150}, {"value": 600}]},
            ],
        }
        out = self.tmp / "t.xlsx"
        _render(td, out)
        wb = load_workbook(str(out))
        ws = wb.active
        merged_ranges = {str(r) for r in ws.merged_cells.ranges}
        self.assertIn("A1:A2", merged_ranges)  # Size row_span=2
        self.assertIn("B1:C1", merged_ranges)  # Rating col_span=2

    def test_cell_types_text_number_fraction_missing_boolean(self):
        td = {
            "schema_version": "pdf2md-table/v1",
            "header_rows": 0,
            "section_dividers": [],
            "continuation_of": None,
            "footnotes": [],
            "rows": [
                {"cells": [
                    {"value": "hello", "type": "text"},
                    {"value": 3.14, "type": "number"},
                    {"value": "1-1/2", "type": "fraction"},
                    {"value": None, "type": "missing", "raw": "—"},
                    {"value": True, "type": "boolean"},
                ]},
            ],
        }
        out = self.tmp / "t.xlsx"
        _render(td, out)
        wb = load_workbook(str(out))
        ws = wb.active
        self.assertEqual(ws.cell(1, 1).value, "hello")
        self.assertEqual(ws.cell(1, 2).value, 3.14)
        self.assertEqual(ws.cell(1, 3).value, "1-1/2")
        self.assertIsNone(ws.cell(1, 4).value)
        self.assertTrue(ws.cell(1, 5).value)

    def test_footnote_attached_as_comment(self):
        td = _td(
            footnotes=[{"marker": "*", "text": "see appendix"}],
        )
        td["rows"][1]["cells"][0]["footnote_markers"] = ["*"]
        out = self.tmp / "t.xlsx"
        _render(td, out)
        wb = load_workbook(str(out))
        ws = wb.active
        cmt = ws.cell(2, 1).comment
        self.assertIsNotNone(cmt)
        self.assertIn("see appendix", cmt.text)


class Determinism(unittest.TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="render_xlsx_"))

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def _hash(self, path: Path) -> str:
        return hashlib.sha256(path.read_bytes()).hexdigest()

    def test_two_runs_byte_identical(self):
        td = _td(footnotes=[{"marker": "*", "text": "note"}])
        a = self.tmp / "a.xlsx"
        b = self.tmp / "b.xlsx"
        _render(td, a)
        _render(td, b)
        self.assertEqual(self._hash(a), self._hash(b))

    def test_complex_table_deterministic(self):
        td = {
            "schema_version": "pdf2md-table/v1",
            "header_rows": 2,
            "section_dividers": [],
            "continuation_of": {"doc_stem": "X", "page": 2, "tbl_ordinal": 1},
            "footnotes": [{"marker": "a", "text": "first"}, {"marker": "b", "text": "second"}],
            "rows": [
                {"cells": [
                    {"value": "Item", "row_span": 2, "is_header": True},
                    {"value": "Vals", "col_span": 2, "is_header": True},
                ]},
                {"cells": [
                    {"value": "min", "is_header": True},
                    {"value": "max", "is_header": True},
                ]},
                {"cells": [
                    {"value": "Pump A", "footnote_markers": ["a"]},
                    {"value": 1.5, "type": "number"},
                    {"value": None, "type": "missing", "footnote_markers": ["b"]},
                ]},
            ],
        }
        a = self.tmp / "a.xlsx"
        b = self.tmp / "b.xlsx"
        _render(td, a)
        _render(td, b)
        self.assertEqual(self._hash(a), self._hash(b))


class CLI(unittest.TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="render_xlsx_"))

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def test_main_writes_file(self):
        td_path = self.tmp / "td.json"
        td_path.write_text(json.dumps(_td()))
        out = self.tmp / "out.xlsx"
        sys.argv = [
            "render_table_xlsx.py",
            "--table-data-json", str(td_path),
            "--output-xlsx", str(out),
            "--caption", "demo",
            "--doc-stem", "DOC",
            "--page", "1",
            "--ordinal", "1",
            "--slug", "demo",
        ]
        rc = _MOD.main()
        self.assertEqual(rc, 0)
        self.assertTrue(out.is_file())

    def test_main_missing_output_parent_returns_2(self):
        td_path = self.tmp / "td.json"
        td_path.write_text(json.dumps(_td()))
        sys.argv = [
            "render_table_xlsx.py",
            "--table-data-json", str(td_path),
            "--output-xlsx", str(self.tmp / "nope" / "out.xlsx"),
            "--caption", "demo",
            "--doc-stem", "DOC",
            "--page", "1",
            "--ordinal", "1",
        ]
        self.assertEqual(_MOD.main(), 2)


if __name__ == "__main__":
    unittest.main(verbosity=2)
